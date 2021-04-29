import random
import configUtilities
from dungeon_utilities import TidyDungeon
from gameMap import GameMap
from loguru import logger
import openpyxl
import time

from metrics import SheetColumns


class RoomsCorridors:

    @staticmethod
    def generate_rooms_and_corridors_map():
        # load config options
        game_config = configUtilities.load_config()

        normal_corridor_width = configUtilities.get_config_value_as_integer(configfile=game_config, section='tunneler',
                                                                            parameter='normal_corridor_width')
        min_step_length = configUtilities.get_config_value_as_integer(configfile=game_config, section='birthOptions',
                                                                      parameter='min_step_length')
        max_step_length = configUtilities.get_config_value_as_integer(configfile=game_config, section='birthOptions',
                                                                      parameter='max_step_length')
        max_age_tunneler = configUtilities.get_config_value_as_integer(configfile=game_config, section='birthOptions',
                                                                       parameter='max_age_tunneler')
        prob_change_dir_step_end = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                               section='tunneler',
                                                                               parameter='prob_change_dir_step_end')

        dungeon_width = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                    parameter='max_width')

        dungeon_height = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                     parameter='max_height')

        game_map = GameMap(mapwidth=dungeon_width, mapheight=dungeon_height)

        sheet_build_id_column = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                            parameter='sheet_build_id')
        sheet_build_start_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                               section='dungeon',
                                                                               parameter='sheet_build_started')
        sheet_build_end_column = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                             parameter='sheet_build_ended')
        sheet_build_duration_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                                  section='dungeon',
                                                                                  parameter='build_duration')
        sheet_build_dungeon_type_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                                  section='dungeon',
                                                                                  parameter='sheet_dungeon_type')
        sheet_build_total_tile_count_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                                  section='dungeon',
                                                                                  parameter='sheet_total_tile_count')
        sheet_build_total_wall_count_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                                  section='dungeon',
                                                                                  parameter='sheet_tiles_as_walls_count')
        sheet_build_total_floor_count_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                                  section='dungeon',
                                                                                  parameter='sheet_tiles_as_floor_count')
        sheet_build_total_corridor_count_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                                  section='dungeon',
                                                                                  parameter='sheet_corridor_count')
        sheet_build_total_room_count_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                                  section='dungeon',
                                                                                  parameter='sheet_room_count')
        sheet_build_total_anteroom_count_column = configUtilities.get_config_value_as_integer(configfile=game_config,
                                                                                  section='dungeon',
                                                                                  parameter='sheet_ante_room_count')


        # New dungeon build starting
        this_workbook = RoomsCorridors.zzz_open_workbook()
        this_worksheet = this_workbook.active
        last_build_id = RoomsCorridors.zzz_get_last_build_id(sheet=this_worksheet)
        this_build_id = last_build_id + 1
        sheet_row_id = 1 + this_build_id
        build_column = SheetColumns.build_id.value
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=build_column,
                                                      row=sheet_row_id, value=this_build_id)
        time_start_string = RoomsCorridors.zzz_get_time_string()
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_start_column,
                                                      row=sheet_row_id, value=time_start_string)

        RoomsCorridors.zzz_save_workbook(workbook=this_workbook)

        tunnel_cur_pos_x = []
        tunnel_cur_pos_y = []
        tunnel_velocity_x = []
        tunnel_velocity_y = []
        tunnel_max_age = []
        tunneler_current_age = []
        tunnel_width = []
        tunnel_change_direction_probability = []

        # setup the first tunneler
        tunnel_cur_pos_x.append(10)
        tunnel_cur_pos_y.append(20)
        tunnel_velocity_x.append(1)
        tunnel_velocity_y.append(0)
        tunnel_max_age.append(max_age_tunneler)
        tunneler_current_age.append(0)
        tunnel_width.append(normal_corridor_width)
        tunnel_change_direction_probability.append(prob_change_dir_step_end)
        tunnel_generate_junction_probability = configUtilities.get_config_value_as_list(configfile=game_config,
                                                                                        section='birthOptions',
                                                                                        parameter='probability_make_junction_room')
        prob_to_spawn_new_tunneler = configUtilities.get_config_value_as_list(configfile=game_config,
                                                                              section='birthOptions',
                                                                              parameter='probability_to_spawn_a_new_tunneler')

        # setup a 2nd tunneler
        # tunnel_cur_pos_x.append(180)
        # tunnel_cur_pos_y.append(20)
        # tunnel_velocity_x.append(-1)
        # tunnel_velocity_y.append(0)
        # tunnel_max_age.append(max_age_tunneler)
        # tunneler_current_age.append(0)
        # tunnel_width.append(normal_corridor_width)
        # tunnel_change_direction_probability.append(prob_change_dir_step_end)

        logger.info('Digging the tunnels')

        digging_tunnels = True
        tunneler_id = 0
        tunneler_count = 0
        current_floor_count = 0
        current_corridor_count = 0
        current_anteroom_count = -1
        current_room_count = -1
        while digging_tunnels:
            age = tunneler_current_age[tunneler_id]
            max_age = tunnel_max_age[tunneler_id]
            digging_tunnels = False
            if age < max_age:
                digging_tunnels = True
                logger.info('working with tunneler {}', tunneler_id)
                current_step_count = 1
                step_length = random.randrange(min_step_length, max_step_length + 1)
                cur_x = tunnel_cur_pos_x[tunneler_id]
                cur_y = tunnel_cur_pos_y[tunneler_id]
                sx = cur_x
                sy = cur_y
                vx = tunnel_velocity_x[tunneler_id]
                vy = tunnel_velocity_y[tunneler_id]
                tunneler_age = tunneler_current_age[tunneler_id]
                max_age_tunneler = tunnel_max_age[tunneler_id]
                tw = tunnel_width[tunneler_id]
                logger.info('[{}] Tunneler age is {}', tunneler_id, tunneler_age)
                logger.info('[{}] And the length of this tunnel is {}', tunneler_id, step_length)

                i_can_dig_this_tunnel = RoomsCorridors.constraint_checker_for_tunnels(startx=cur_x, starty=cur_y,
                                                                                      tunnel_length=step_length)
                # carve tunnel for max step length
                if i_can_dig_this_tunnel:
                    current_corridor_count += 1
                    if age > 9:
                        tunneler_age = 9
                    else:
                        tunneler_age = age
                    logger.info('[{}] It appears there is room to carve out this part of the tunnel', tunneler_id)
                    prob_to_spawn_tunneler = -1
                    # prob_to_spawn_tunneler = int(prob_to_spawn_new_tunneler[tunneler_age])
                    prob_to_spawn_room = 25
                    while current_step_count <= step_length:
                        cur_x += vx
                        cur_y += vy
                        tunnel_floor_count = RoomsCorridors.carve_tunnel(game_map=game_map, vx=vx, vy=vy, cur_x=cur_x, cur_y=cur_y,
                                                    tunnel_width=tw)
                        current_floor_count += tunnel_floor_count
                        tunnel_cur_pos_x[tunneler_id] += vx
                        tunnel_cur_pos_y[tunneler_id] += vy
                        current_step_count += 1
                        # check if we want to spawn a new tunneler
                        # and increase tunneler_count
                        # for every iteration within the step determine if a new tunneler should be spawned
                        rr = random.randrange(0, 100)
                        if prob_to_spawn_tunneler > rr:
                            # spawn a new tunneler
                            logger.info('=== NEW TUNNELER SPAWNED ===')
                            tunneler_count += 1
                            logger.info('With an id of {}', tunneler_count)
                            if vx != 0:
                                logger.info('[{}] will be carving south', tunneler_count)
                                tunnel_cur_pos_x.append(tunnel_cur_pos_x[tunneler_id])
                                tunnel_cur_pos_y.append(tunnel_cur_pos_y[tunneler_id] - 1)
                                tunnel_velocity_x.append(0)
                                tunnel_velocity_y.append(-1)
                            if vy != 0:
                                logger.info('[{}] will be carving east', tunneler_count)
                                tunnel_cur_pos_x.append(tunnel_cur_pos_x[tunneler_id] + 1)
                                tunnel_cur_pos_y.append(tunnel_cur_pos_y[tunneler_id])
                                tunnel_velocity_x.append(1)
                                tunnel_velocity_y.append(0)
                            tunnel_max_age.append(max_age_tunneler)
                            tunneler_current_age.append(0)
                            tunnel_width.append(normal_corridor_width)
                            tunnel_change_direction_probability.append(prob_change_dir_step_end)
                    if vy != 0:
                        logger.info('[{}] Wants to carve out a side-room tunnel', tunneler_id)
                        # room_chance = random.randrange(0, 100)
                        room_chance = 1000
                        if room_chance > prob_to_spawn_room:
                            RoomsCorridors.add_room_to_tunnel(game_map=game_map, cvx=vx, cvy=vy, start_pos=[sx, sy], end_pos=[cur_x, cur_y])

                else:
                    logger.info('[{}] There is not enough room to carve out this part of the tunnel', tunneler_id)
                vx, vy, dir_changed = RoomsCorridors.choose_random_direction(curx=cur_x, cury=cur_y, cvx=vx, cvy=vy,
                                                                             prob_change_dir_step_end=prob_change_dir_step_end,
                                                                             step_length=step_length)
                # dir_changed = True
                # vx = 0
                # vy = 1
                junction_created = False
                if dir_changed:
                    # check for chance to add a junction room
                    # newx, newy, junction_created, junction_floor_count = RoomsCorridors.create_junction_yes_or_no(
                    #     velocity=[tunnel_velocity_x[tunneler_id], tunnel_velocity_y[tunneler_id]], game_map=game_map,
                    #     probability_list=tunnel_generate_junction_probability, new_velocity=[vx, vy],
                    #     position=[cur_x, cur_y], tunnel_width=tw)
                    if junction_created:
                        current_anteroom_count += 1
                        tunnel_cur_pos_x[tunneler_id] = newx
                        tunnel_cur_pos_y[tunneler_id] = newy
                        current_floor_count += junction_floor_count
                # shall we change tunnel widths
                change_tunnel_width, new_tunnel_width = RoomsCorridors.should_tunnel_width_change(
                    tunnel_age=tunneler_age, junction_created=junction_created, current_tunnel_width=tw)
                if change_tunnel_width:
                    tunnel_width[tunneler_id] = new_tunnel_width
                tunnel_velocity_x[tunneler_id] = vx
                tunnel_velocity_y[tunneler_id] = vy
                # has the current tunneler reached its' end of life
                if tunneler_age >= max_age_tunneler:
                    logger.info('[{}] has reached its maximum age', tunneler_id)
                else:
                    age = tunneler_current_age[tunneler_id]
                    age += 1
                    tunneler_current_age[tunneler_id] = age
                    logger.info('[{}]has aged', tunneler_id)
            tunneler_id += 1
            if tunneler_id > 0:
                tunneler_id = 0
        decorate_map = TidyDungeon(game_map=game_map)
        total_wall_count = decorate_map.add_walls()

        # dungeon build has ended
        # capture time ended for this dungeon build
        time_end_string = RoomsCorridors.zzz_get_time_string()
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_end_column,
                                                      row=sheet_row_id, value=time_end_string)

        # calculate the duration of this dungeon build
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_duration_column,
                                                      row=sheet_row_id,
                                                      value="=sum(C" + str(sheet_row_id) + "-" + "B" + str(
                                                          sheet_row_id) + ")")
        # and format the cell as HH:MM:SS
        RoomsCorridors.zzz_set_cell_to_time_format(sheet=this_worksheet, column=sheet_build_duration_column,
                                                   row=sheet_row_id)
        # dungeon type being built
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_dungeon_type_column,
                                                      row=sheet_row_id, value="tunnelling")
        # total tile count for this dungeon
        tile_count = dungeon_width * dungeon_height
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_total_tile_count_column,
                                                      row=sheet_row_id, value=tile_count)

        # number of tiles marked as wall
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_total_wall_count_column,
                                                      row=sheet_row_id, value=total_wall_count)

        # number of tiles marked as floor
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_total_floor_count_column,
                                                      row=sheet_row_id, value=current_floor_count)
        # number of corridors
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_total_corridor_count_column,
                                                      row=sheet_row_id, value=current_corridor_count)
        # number of rooms
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_total_room_count_column,
                                                      row=sheet_row_id, value=current_room_count)
        # number of ante-rooms
        RoomsCorridors.zzz_write_value_to_spreadsheet(sheet=this_worksheet, column=sheet_build_total_anteroom_count_column,
                                                      row=sheet_row_id, value=current_anteroom_count)

        # save the dungeon metrics
        RoomsCorridors.zzz_save_workbook(workbook=this_workbook)

        return game_map

    @staticmethod
    def constraint_checker_for_tunnels(startx, starty, tunnel_length):
        # load config options
        game_config = configUtilities.load_config()
        dungeon_width = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                    parameter='max_width')

        dungeon_height = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                     parameter='max_height')

        dw = dungeon_width - 2
        dh = dungeon_height - 3

        can_dig_that_tunnel = False
        if (startx - tunnel_length > 2) and (startx + tunnel_length < dw) and (starty - tunnel_length > 2) and (
                starty + tunnel_length < dh):
            can_dig_that_tunnel = True

        return can_dig_that_tunnel

    @staticmethod
    def should_tunnel_width_change(tunnel_age, junction_created, current_tunnel_width):
        # load config options
        game_config = configUtilities.load_config()
        change_tunnel_width = False
        new_tunnel_width = 0
        if junction_created:
            tunnel_change_width_probability = configUtilities.get_config_value_as_list(configfile=game_config,
                                                                                       section='birthOptions',
                                                                                       parameter='tw_change_probability_junction')
        else:
            tunnel_change_width_probability = configUtilities.get_config_value_as_list(configfile=game_config,
                                                                                       section='birthOptions',
                                                                                       parameter='tw_change_probability_no_junction')

        if tunnel_age > 9:
            tunneler_age = 9
        else:
            tunneler_age = tunnel_age
        change_width_probability = int(tunnel_change_width_probability[tunneler_age])
        if random.randrange(0, 100) < change_width_probability:
            change_tunnel_width = True
            if current_tunnel_width == 1 or current_tunnel_width == 5:
                new_tunnel_width = 3

            if current_tunnel_width == 3:
                coin_flip = random.randrange(0, 100)
                if coin_flip > 75:
                    new_tunnel_width = 5
                else:
                    new_tunnel_width = 1
        return change_tunnel_width, new_tunnel_width

    @staticmethod
    def create_junction_yes_or_no(game_map, probability_list, new_velocity, position, velocity, tunnel_width):
        junction_created = False
        newx = 0
        newy = 0
        # probability_percentage = int(probability_list[tunneler_age])
        probability_percentage = 1
        if probability_percentage > 0:
            create_room_percentage = random.randrange(50, 100)
            if create_room_percentage > probability_percentage:
                # create junction room
                newx, newy, junction_created, floor_count = RoomsCorridors.carve_out_junction(game_map=game_map, position=position,
                                                                                 tunnel_width=tunnel_width,
                                                                                 velocity=velocity,
                                                                                 new_velocity=new_velocity)

        return newx, newy, junction_created, floor_count

    @staticmethod
    def carve_out_junction(game_map, position, velocity, tunnel_width, new_velocity):
        game_config = configUtilities.load_config()
        floor_count = 0
        curx = position[0]
        cury = position[1]
        newx = curx
        newy = cury
        logger.info('curx is {} and cury is {}', curx, cury)
        # in this context vx/vy represent the direction the tunnel is being carved out of
        vx = velocity[0]
        vy = velocity[1]
        logger.info('current velocity is {}/{}', vx, vy)
        # in this context new_velocity represents the direction the next tunnel will be travelling in
        new_vx = new_velocity[0]
        new_vy = new_velocity[1]
        junction_carved = False
        left_tile = 0
        right_tile = 0
        top_tile = 0
        bottom_tile = 0
        chosen_tunnel_width = RoomsCorridors.choose_which_size_junction_to_carve(game_config=game_config)
        # logger.info('new tunnel width is {}', chosen_tunnel_width)
        # in this context vx/vy represent the direction the tunnel has just been carved out of
        if vy > 0:
            if chosen_tunnel_width == 1:
                # 3 x 3
                left_tile = curx - 1
                right_tile = curx + 2
                bottom_tile = cury + 4
                newy = bottom_tile - 2

            if chosen_tunnel_width == 2:
                # 5 x 5
                left_tile = curx - 2
                right_tile = curx + 3
                bottom_tile = cury + 6
                newy = bottom_tile - 3

            if chosen_tunnel_width == 3:
                # 7 x 5
                left_tile = curx - 3
                right_tile = curx + 4
                bottom_tile = cury + 6
                newy = bottom_tile - 3

            top_tile = cury + 1
            if new_vx == -1:
                newx = left_tile
            if new_vx == 1:
                newx = right_tile - 1

        if vy < 0:
            # 3 x 3
            if chosen_tunnel_width == 1:
                left_tile = curx - 1
                right_tile = curx + 2
                top_tile = cury - 3
                newy = top_tile + 1

            if chosen_tunnel_width == 2:
                # 5 x 5
                left_tile = curx - 2
                right_tile = curx + 3
                top_tile = cury - 5
                newy = top_tile + 2

            if chosen_tunnel_width == 3:
                # 7 x 5
                left_tile = curx - 3
                right_tile = curx + 4
                top_tile = cury - 5
                newy = top_tile + 2

            bottom_tile = cury
            if new_vx == 1:
                newx = right_tile - 1
            if new_vx == -1:
                newx = left_tile

        if vx > 0:
            if chosen_tunnel_width == 1:
                # 3 x 3
                top_tile = cury - 1
                bottom_tile = cury + 2
                right_tile = curx + 4

            if chosen_tunnel_width == 2:
                # 5 x 4
                top_tile = cury - 2
                bottom_tile = cury + 3
                right_tile = curx + 6

            if chosen_tunnel_width == 3:
                # 7 x 5
                top_tile = cury - 3
                bottom_tile = cury + 4
                right_tile = curx + 5

            left_tile = curx + 1
            # this code looks to 'centre' the position of the next tunneler
            if new_vy == 1:
                newx = right_tile - 2
                newy = bottom_tile - 1
            if new_vy == -1:
                newx = right_tile - 2
                newy = top_tile
            if new_vx == 1:
                newx = right_tile - 1

        if vx < 0:
            if chosen_tunnel_width == 1:
                # 3 x 3
                top_tile = cury - 1
                bottom_tile = cury + 2
                left_tile = curx - 3
                newx = left_tile + 1

            if chosen_tunnel_width == 2:
                # 5 x 4
                top_tile = cury - 2
                bottom_tile = cury + 3
                left_tile = curx - 5
                newx = left_tile + 2

            if chosen_tunnel_width == 3:
                # 7 x 5
                top_tile = cury - 3
                bottom_tile = cury + 4
                left_tile = curx - 5
                newx = left_tile + 2

            right_tile = curx
            if new_vy == -1:
                newx = left_tile
            if new_vy == 1:
                newy = bottom_tile - 1
        # logger.info('top tile is {} and bottom tile is {}', top_tile, bottom_tile)
        # logger.info('left tile is {} and right tile is {}', left_tile, right_tile)
        if (top_tile > 2 and bottom_tile < 58) and (left_tile > 2 and right_tile < 158):
            floor_count = RoomsCorridors.carve_out_floor_tiles(game_map=game_map, top_tile=top_tile, bottom_tile=bottom_tile, left_tile=left_tile, right_tile=right_tile)
            junction_carved = True
        return newx, newy, junction_carved, floor_count

    @staticmethod
    def choose_which_size_junction_to_carve(game_config):
        # load config options
        junction_size_normal_width_list = configUtilities.get_config_value_as_list(configfile=game_config,
                                                                                   section='birthOptions',
                                                                                   parameter='probability_to_generate_room_based_on_normal_width_tunnel')

        probability_for_normal_width = int(junction_size_normal_width_list[0])
        probability_for_wide_width = int(junction_size_normal_width_list[1])
        probability_for_wider_width = int(junction_size_normal_width_list[2])

        choose_tunnel_width_percentage = random.randrange(0, 100)
        chosen_tunnel_width = 0
        if probability_for_normal_width > 0:
            if choose_tunnel_width_percentage < probability_for_normal_width:
                chosen_tunnel_width = 1

        if probability_for_wide_width > 0:
            if choose_tunnel_width_percentage > probability_for_normal_width and choose_tunnel_width_percentage > probability_for_wide_width:
                chosen_tunnel_width = 2

        if probability_for_wider_width > 0:
            if choose_tunnel_width_percentage < probability_for_wider_width:
                chosen_tunnel_width = 3

        return chosen_tunnel_width

    @staticmethod
    def carve_tunnel(game_map, vx, vy, cur_x, cur_y, tunnel_width):
        # load config options
        game_config = configUtilities.load_config()
        tile_type_floor = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                      parameter='TILE_TYPE_FLOOR')
        dungeon_width = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                    parameter='max_width')

        dungeon_height = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                     parameter='max_height')

        floor_count = 0

        # carve tunnel - walkable space + corridor edges
        # walkable space = tunnel_width
        # central corridor
        game_map.tiles[cur_x][cur_y].type_of_tile = tile_type_floor
        game_map.tiles[cur_x][cur_y].blocked = False
        floor_count += 1
        # corridor width
        tw = 4
        if tunnel_width == 1:
            tw = 0
        if tunnel_width == 3:
            tw = 2

        if tunnel_width == 5:
            tw = 3

        for a in range(1, tw):
            if vy != 0:
                left_w = cur_x - a
                right_w = cur_x + a
                # leftside of central path
                if left_w > 1:
                    game_map.tiles[left_w][cur_y].type_of_tile = tile_type_floor
                    game_map.tiles[left_w][cur_y].blocked = False
                    floor_count += 1
                if right_w < dungeon_width:
                    game_map.tiles[right_w][cur_y].type_of_tile = tile_type_floor
                    game_map.tiles[right_w][cur_y].blocked = False
                    floor_count += 1

            if vx != 0:
                top_w = cur_y - a
                bot_w = cur_y + a
                # rightside of central path
                if top_w > 1:
                    game_map.tiles[cur_x][top_w].type_of_tile = tile_type_floor
                    game_map.tiles[cur_x][top_w].blocked = False
                    floor_count += 1
                if bot_w < dungeon_height:
                    game_map.tiles[cur_x][bot_w].type_of_tile = tile_type_floor
                    game_map.tiles[cur_x][bot_w].blocked = False
                    floor_count += 1

        return floor_count

    @staticmethod
    def choose_random_direction(curx, cury, cvx, cvy, prob_change_dir_step_end, step_length):
        vx = cvx
        vy = cvy
        last_direction = 0
        dir_changed = False
        last_dir_string = ''
        if vx != 0 and vy == 0:
            # last carved out an east/west tunnel
            last_direction = 1
            last_dir_string = 'east/west'
        if vy != 0 and vx == 0:
            # last carved out a north/south tunnel
            last_direction = 2
            last_dir_string = 'north/south'
        if last_direction > 0:
            logger.info('Tunnel was traveling {}', last_dir_string)
        direction_changed = False
        change_dir = random.randrange(0, 100)
        if change_dir <= prob_change_dir_step_end:
            while not direction_changed:
                vx, vy, direction_changed = RoomsCorridors.check_direction_to_carve(curx=curx, cury=cury,
                                                                                    last_direction=last_direction,
                                                                                    step_length=step_length)
                dir_changed = True
            logger.info('Velocity parms reset: vx={} and vy={}', vx, vy)
        else:
            logger.info('No change in tunnel direction')
        return vx, vy, dir_changed

    @staticmethod
    def check_direction_to_carve(curx, cury, last_direction, step_length):
        step_length += 1
        vx = 0
        vy = 0
        # when last_direction is 1 - east or west
        # last_direction == 2 - north or south
        direction_changed = False
        if last_direction != 2:
            rand_direction = random.randrange(1, 3)
            # carve north
            if rand_direction == 1 and cury - step_length > 1:
                logger.info('North/South direction change set to North')
                # move north
                vx = 0
                vy = -1
                direction_changed = True
            # carve south
            if rand_direction == 2 and cury + step_length < 50:
                logger.info('North/South direction change set to South')
                # move south
                vx = 0
                vy = 1
                direction_changed = True
        if last_direction != 1:
            rand_direction = random.randrange(1, 3)
            if rand_direction == 1 and curx + step_length < 198:
                logger.info('East/West direction change set to East')
                # move east
                vx = 1
                vy = 0
                direction_changed = True
            if rand_direction == 2 and curx - step_length > 1:
                logger.info('East/West direction change set to West')
                # move west
                vx = -1
                vy = 0
                direction_changed = True
        logger.info('vx is now {} and vy is now {}', vx, vy)
        return vx, vy, direction_changed

    @staticmethod
    def zzz_open_workbook():
        workbook = openpyxl.load_workbook(filename="dungeon_metrics.xlsx")
        return workbook

    @staticmethod
    def zzz_write_value_to_spreadsheet(sheet, column, row, value):
        sheet.cell(row=row, column=column, value=value)

    @staticmethod
    def zzz_set_cell_to_time_format(sheet, column, row):
        cell = sheet.cell(column=column, row=row)
        cell.number_format = "HH:MM:SS"

    @staticmethod
    def zzz_save_workbook(workbook):
        workbook.save(filename="dungeon_metrics.xlsx")

    @staticmethod
    def zzz_get_time_string():
        return time.strftime("%H:%M:%S", time.localtime())

    @staticmethod
    def zzz_get_last_build_id(sheet):
        last_build = 0
        for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
            for r in row:
                if isinstance(r, int) and r > 0:
                    last_build = r
        return last_build

    @staticmethod
    def add_room_to_tunnel(game_map, cvx, cvy, start_pos, end_pos):
        # cvx & cvy let me know the direction of the tunnel

        # these hold the start and end coords of the tunnel
        # and allow for a room to be generated anywhere along that tunnel length
        sx = start_pos[0]
        sy = start_pos[1]
        ex = end_pos[0]
        ey = end_pos[1]

        logger.debug('new room being carved out')

        top_tile = 0
        bottom_tile = 0
        left_tile = 0
        right_tile = 0

        # these are the dimensions of the room
        room_width = 9
        room_height = 9

        # is tunnel travelling horizontally
        if cvx != 0:
            top_tile = 0
            bottom_tile = 0
            left_tile = 0
            right_tile = 0

        # is tunnel travelling vertically
        if cvy != 0:
            top_tile = sy - 5
            bottom_tile = sy + 4
            left_tile = sx - room_width
            right_tile = sx - 2

        # this code carves out a rectangle of floor tiles
        floor_count = RoomsCorridors.carve_out_floor_tiles(game_map=game_map, top_tile=top_tile,
                                                           bottom_tile=bottom_tile, left_tile=left_tile, right_tile=right_tile)
        return floor_count

    @staticmethod
    def carve_out_floor_tiles(game_map, top_tile, bottom_tile, left_tile, right_tile):
        # load config options
        game_config = configUtilities.load_config()
        tile_type_floor = configUtilities.get_config_value_as_integer(configfile=game_config, section='dungeon',
                                                                      parameter='TILE_TYPE_FLOOR')
        floor_count = 0

        for yy in range(top_tile, bottom_tile):
            for xx in range(left_tile, right_tile):
                game_map.tiles[xx][yy].type_of_tile = tile_type_floor
                game_map.tiles[xx][yy].blocked = False
                floor_count += 1

        return floor_count
