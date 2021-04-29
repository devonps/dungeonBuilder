import enum
from datetime import datetime

import openpyxl
from loguru import logger


class SheetColumns(enum.Enum):
    build_id = enum.auto()
    build_seed = enum.auto()
    build_started = enum.auto()
    build_ended = enum.auto()
    build_duration = enum.auto()
    build_type = enum.auto()
    total_tile_count = enum.auto()
    total_wall_count = enum.auto()
    total_obstacle_count = enum.auto()
    total_floor_count = enum.auto()
    total_corridor_count = enum.auto()
    total_room_count = enum.auto()
    total_ante_room_count = enum.auto()


class Metrics:

    @staticmethod
    def initialise():
        current_workbook = Metrics.workbook_open()
        current_worksheet = current_workbook.active
        return current_workbook, current_worksheet

    @staticmethod
    def update_build_id(worksheet):
        last_build_id = Metrics.get_last_build_id(worksheet=worksheet)
        this_build_id = last_build_id + 1
        return this_build_id

    @staticmethod
    def set_current_sheet_row_id(this_build_id):
        return this_build_id + 1

    @staticmethod
    def get_last_build_id(worksheet):
        last_build = 0
        for row in worksheet.iter_rows(min_row=1, max_col=1, values_only=True):
            for r in row:
                if isinstance(r, int) and r > 0:
                    last_build = r
        return last_build

    @staticmethod
    def workbook_open():
        workbook = openpyxl.load_workbook(filename="dungeon_metrics.xlsx")
        return workbook

    @staticmethod
    def update_worksheet(sheet, column, row, value):
        sheet.cell(row=row, column=column, value=value)

    @staticmethod
    def save_workbook(workbook):
        workbook.save(filename="dungeon_metrics.xlsx")

    @staticmethod
    def get_time_string():
        time = datetime.now()
        dt = datetime.time(time)
        logger.info('Time stamp is {}', dt)
        return dt

    @staticmethod
    def set_cell_to_time_format(sheet, column, row):
        cell = sheet.cell(column=column, row=row)
        cell.number_format = "HH:MM:SS.0000"

    @staticmethod
    def set_build_seed_cell(sheet, column, row):
        cell = sheet.cell(column=column, row=row)
        cell.number_format = "0"
